"""
Export Routes
GET /api/export/excel   — full Excel with filters
GET /api/export/pdf     — PDF report with filters
"""
from __future__ import annotations
import io
from datetime import datetime
from flask import Blueprint, request, send_file
from flask_jwt_extended import jwt_required
from sqlalchemy import and_
from models.database import MilkRecord

export_bp = Blueprint("export", __name__)


def _build_query(args: dict):
    filters = []
    if args.get("decision"):
        filters.append(MilkRecord.decision == args["decision"])
    if args.get("fraud_risk"):
        filters.append(MilkRecord.fraud_risk == args["fraud_risk"])
    if args.get("shift"):
        filters.append(MilkRecord.shift == args["shift"])
    if args.get("date_from"):
        try:
            filters.append(MilkRecord.date >= datetime.strptime(args["date_from"], "%Y-%m-%d").date())
        except ValueError:
            pass
    if args.get("date_to"):
        try:
            filters.append(MilkRecord.date <= datetime.strptime(args["date_to"], "%Y-%m-%d").date())
        except ValueError:
            pass
    if args.get("farmer_code"):
        filters.append(MilkRecord.farmer_code == args["farmer_code"])
    q = MilkRecord.query
    if filters:
        q = q.filter(and_(*filters))
    return q.order_by(MilkRecord.date.desc())


@export_bp.get("/excel")
@jwt_required()
def export_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    args = dict(request.args)
    records = _build_query(args).limit(10_000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Milk Quality Records"

    header_fill = PatternFill("solid", fgColor="1A3C5E")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ID","Farmer Name","Farmer Code","Date","Shift","FAT %","SNF %","pH","Acidity",
               "Temp (°C)","Sp. Gravity","COB Test","Alcohol Test","Organoleptic","Sediment",
               "MBRT (h)","Raw Temp","Qty (L)","Decision","Fraud Risk","Reasons"]

    ws.append(headers)
    ws.row_dimensions[1].height = 22
    for c in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr

    dec_fills = {
        "accept":       PatternFill("solid", fgColor="C8E6C9"),
        "reject":       PatternFill("solid", fgColor="FFCDD2"),
    }

    def _f(v, fmt="{:.2f}"):
        return fmt.format(float(v)) if v is not None else ""

    for rec in records:
        ws.append([
            rec.id, rec.farmer_name, rec.farmer_code or "", str(rec.date),
            (rec.shift or "").capitalize(),
            _f(rec.fat), _f(rec.snf), _f(rec.ph), _f(rec.acidity, "{:.3f}"),
            _f(rec.temperature, "{:.1f}"), _f(rec.specific_gravity, "{:.4f}"),
            rec.cob_test, rec.alcohol_test, rec.organoleptic, rec.sediment_test,
            _f(rec.mbrt, "{:.1f}"), _f(rec.raw_milk_temp, "{:.1f}"),
            _f(rec.quantity, "{:.2f}"),
            (rec.decision or "").upper(),
            (rec.fraud_risk or "").upper(),
            "; ".join(rec.reasons or []),
        ])
        ri = ws.max_row
        for c in range(1, len(headers)+1):
            ws.cell(row=ri, column=c).border = bdr
        dec_cell = ws.cell(row=ri, column=19)
        df = dec_fills.get(rec.decision)
        if df:
            dec_cell.fill = df
            dec_cell.font = Font(bold=True)

    col_widths = [6,22,14,12,10,8,8,7,9,10,14,12,14,14,12,10,12,10,16,14,70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.append(["Total Records", len(records)])
    ws2.append(["Accepted",  sum(1 for r in records if r.decision == "accept")])
    ws2.append(["Rejected",  sum(1 for r in records if r.decision == "reject")])
    ws2.append(["Fraud High", sum(1 for r in records if r.fraud_risk == "high")])
    ws2.append(["Fraud Medium", sum(1 for r in records if r.fraud_risk == "medium")])
    ws2.append(["Generated", datetime.now().strftime("%d %b %Y %H:%M")])
    for cell in ws2["A"]:
        cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"milk_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=fname)


@export_bp.get("/pdf")
@jwt_required()
def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    args = dict(request.args)
    records = _build_query(args).limit(500).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=12*mm, rightMargin=12*mm,
                             topMargin=14*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=styles["Heading1"],
                              textColor=colors.HexColor("#1A3C5E"),
                              fontSize=14, spaceAfter=3, fontName="Helvetica-Bold")
    sub_s   = ParagraphStyle("S", parent=styles["Normal"],
                              textColor=colors.HexColor("#64748B"),
                              fontSize=8, spaceAfter=0)
    elements = []
    elements.append(Paragraph("Smart Milk Decision Tool System — Quality Report", title_s))
    acc = sum(1 for r in records if r.decision=="accept")
    rej = sum(1 for r in records if r.decision=="reject")
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  "
        f"Total: {len(records)}  |  Accepted: {acc}  |  Rejected: {rej}",
        sub_s))
    elements.append(Spacer(1, 4*mm))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1")))
    elements.append(Spacer(1, 4*mm))

    header = ["ID","Farmer","Code","Date","Shift","FAT","SNF","pH","Acid.","Temp",
              "Sp.G","MBRT","COB","Alc","Org","Sed","Decision","Fraud"]
    col_w = [10*mm,32*mm,18*mm,20*mm,14*mm,11*mm,11*mm,10*mm,11*mm,12*mm,
             14*mm,11*mm,10*mm,10*mm,10*mm,10*mm,22*mm,14*mm]

    def _v(v, fmt="{:.2f}"):
        return fmt.format(float(v)) if v is not None else "—"

    table_data = [header]
    for rec in records:
        table_data.append([
            str(rec.id), (rec.farmer_name or "")[:18], rec.farmer_code or "—",
            str(rec.date), (rec.shift or "").capitalize(),
            _v(rec.fat), _v(rec.snf), _v(rec.ph), _v(rec.acidity, "{:.3f}"),
            _v(rec.temperature, "{:.1f}"), _v(rec.specific_gravity, "{:.4f}"),
            _v(rec.mbrt, "{:.1f}"),
            "POS" if rec.cob_test == "positive" else "NEG",
            "POS" if rec.alcohol_test == "positive" else "NEG",
            "ABN" if rec.organoleptic == "abnormal" else "NOR",
            "DTY" if rec.sediment_test == "dirty" else "CLN",
            (rec.decision or "").upper().replace("_"," "),
            (rec.fraud_risk or "").upper(),
        ])

    tbl = Table(table_data, colWidths=col_w, repeatRows=1)
    ts = [
        ("BACKGROUND", (0,0),(-1,0), colors.HexColor("#1A3C5E")),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0),(-1,-1), 7),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",       (0,0),(-1,-1), 0.4, colors.HexColor("#E2E8F0")),
        ("ALIGN",      (0,0),(-1,-1), "CENTER"),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
    ]
    for i, rec in enumerate(records, 1):
        dc = 16
        if rec.decision == "accept":
            ts += [("BACKGROUND",(dc,i),(dc,i),colors.HexColor("#D1FAE5")),
                   ("TEXTCOLOR",(dc,i),(dc,i),colors.HexColor("#065F46")),
                   ("FONTNAME",(dc,i),(dc,i),"Helvetica-Bold")]
        elif rec.decision == "reject":
            ts += [("BACKGROUND",(dc,i),(dc,i),colors.HexColor("#FEE2E2")),
                   ("TEXTCOLOR",(dc,i),(dc,i),colors.HexColor("#991B1B")),
                   ("FONTNAME",(dc,i),(dc,i),"Helvetica-Bold")]
        if rec.fraud_risk == "high":
            ts += [("BACKGROUND",(17,i),(17,i),colors.HexColor("#FEE2E2")),
                   ("TEXTCOLOR",(17,i),(17,i),colors.HexColor("#991B1B")),
                   ("FONTNAME",(17,i),(17,i),"Helvetica-Bold")]
        if rec.cob_test == "positive":
            ts += [("TEXTCOLOR",(12,i),(12,i),colors.HexColor("#DC2626")),
                   ("FONTNAME",(12,i),(12,i),"Helvetica-Bold")]
        if rec.alcohol_test == "positive":
            ts += [("TEXTCOLOR",(13,i),(13,i),colors.HexColor("#DC2626")),
                   ("FONTNAME",(13,i),(13,i),"Helvetica-Bold")]
    tbl.setStyle(TableStyle(ts))
    elements.append(tbl)
    doc.build(elements)
    buf.seek(0)
    fname = f"milk_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=fname)
